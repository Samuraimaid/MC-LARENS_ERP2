from __future__ import annotations

from typing import Any, Tuple

from fastapi import HTTPException


def get_openpyxl_symbols() -> Tuple[Any, Any]:
    try:
        from openpyxl import Workbook as _Workbook, load_workbook as _load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="Excel export/import dependencies are not installed",
        ) from exc
    return _Workbook, _load_workbook


def get_reportlab_symbols() -> Tuple[Any, Any, Any, Any]:
    try:
        from reportlab.lib import colors as _colors
        from reportlab.lib.pagesizes import letter as _letter
        from reportlab.lib.utils import ImageReader as _ImageReader
        from reportlab.pdfgen import canvas as _canvas
    except ImportError as exc:
        raise HTTPException(
            status_code=503,
            detail="PDF export dependencies are not installed",
        ) from exc
    return _colors, _letter, _ImageReader, _canvas
