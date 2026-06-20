import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def safe_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", name).strip()
    cleaned = cleaned[:31] or "Hoja"
    base = cleaned
    i = 1
    while cleaned in existing:
        suffix = f"_{i}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1
    return cleaned


def extract_year(descriptor: str) -> str:
    if not descriptor:
        return "N/D"
    match = re.search(r"\[([^\]]+)\]", descriptor)
    return match.group(1).strip() if match else "N/D"


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    source = root / "frontend" / "src" / "data" / "vehicleCatalog.json"
    output = root / "vehicle_catalog_by_brand.xlsx"

    data = json.loads(source.read_text(encoding="utf-8"))
    entries = data.get("entries", [])

    grouped: dict[str, list[dict]] = {}
    for entry in entries:
        brand = (entry.get("brand") or "SIN MARCA").strip() or "SIN MARCA"
        grouped.setdefault(brand, []).append(entry)

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    used_sheet_names: set[str] = set()

    for brand in sorted(grouped.keys()):
        sheet_name = safe_sheet_name(brand, used_sheet_names)
        used_sheet_names.add(sheet_name)
        ws = workbook.create_sheet(title=sheet_name)

        ws.append(["Año", "Modelo", "Variación", "Descriptor"])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        rows = []
        for e in grouped[brand]:
            year = extract_year(e.get("descriptor", ""))
            model = (e.get("model") or "").strip() or "N/D"
            variation = (e.get("engine") or "").strip() or "N/D"
            descriptor = (e.get("descriptor") or "").strip() or "N/D"
            rows.append((year, model, variation, descriptor))

        rows.sort(key=lambda x: (x[0], x[1], x[2], x[3]))
        for r in rows:
            ws.append(list(r))

        widths = {1: 16, 2: 30, 3: 28, 4: 42}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

    workbook.save(output)
    print(f"Archivo generado: {output}")
    print(f"Marcas (hojas): {len(grouped)}")
    print(f"Filas de variaciones: {len(entries)}")


if __name__ == "__main__":
    main()
